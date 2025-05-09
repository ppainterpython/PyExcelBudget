# ---------------------------------------------------------------------------- +
#region PyExeclBudget - read Excel files from a bank and analyze budget.
""" p3ExcelBudget - read Excel files from a bank and analyze budget.

    This module is the main entry point for the p3ExcelBudget application.
    It reads Excel files from a bank and analyzes the budget.
"""
#endregion PyExeclBudget - read Excel files from a bank and analyze budget.
# ---------------------------------------------------------------------------- +
#region Imports
# python standard libraries packages and modules 
import atexit, pathlib, logging, inspect, logging.config  #, logging.handlers

# third-party  packages and module libraries
from openpyxl import Workbook
import inspect, pyjson5
from config import settings
import p3logging as p3l, p3_utils as p3u

# local packages and module libraries
import view_model.bdm_view_model as p3bmvm
import view.budmod_cli_view as p3bmv
logger = logging.getLogger(settings.app_name)
logger.propagate = True
#endregion Imports
# ---------------------------------------------------------------------------- +
#region configure_logging() function
def configure_logging(logger_name : str = settings.app_name, logtest : bool = False) -> None:
    """Setup the application logger."""
    try:
        # Configure logging
        log_config_file = "budget_model_logging_config.jsonc"
        # Set the log filename for this application.
        # filenames = {"file": "logs/p3ExcelBudget.log"}
        _ = p3l.setup_logging(
            logger_name = logger_name,
            config_file = log_config_file
            )
        p3l.set_log_flag(p3l.LOG_FLAG_PRINT_CONFIG_ERRORS, True)
        logger = logging.getLogger(settings.app_name)
        logger.setLevel(logging.DEBUG)
        logger.info("+ ----------------------------------------------------- +")
        logger.info(f"+ Running {settings.app_name}...")
        logger.info("+ ----------------------------------------------------- +")
        if(logtest): 
            p3l.quick_logging_test(settings.app_name, log_config_file, reload = False)
    except Exception as e:
        logger.error(p3u.exc_msg(configure_logging, e))
        raise
#endregion configure_logging() function
# ---------------------------------------------------------------------------- +
#region log_workbook_info() function
def log_workbook_info(file_name : str = "unknown",wb : Workbook = None) -> None:  
    if wb is None or not isinstance(wb, Workbook):
        logger.error(f"Workbook({file_name}): None or not a Workbook.")
        return
    logger.info(f"Workbook({file_name}): with sheets: {str(wb.sheetnames)}")
    # sheet = wb.active
    # r = sheet.max_row
    # c = sheet.max_column
    logger.info(f"  Active sheet({wb.active.title}): " 
                f"({wb.active.max_row} x {wb.active.max_column})")
#endregion log_workbook_info() function
# ---------------------------------------------------------------------------- +
#region budmod() function
def budmod():
    """Main function to run PyExcelBudget application."""
    try:
        # p3bm.tryout_budget_model_template()

        trans_file = "BOAChecking2025.xlsx"
        # Initalize the ViewModel and View
        # bmt = p3bm.BudgetModelTemplate()
        # bm = p3bm.BudgetModel().bdm_initialize(bmt) # use the template to init
        # p3bm.log_BDM_info(bm)
        bmvm = p3bmvm.BudgetModelCommandViewModel()
        bmvm.initialize() # Initialize the BudgetModelCommandViewModel
        p3u.set_print_output(True)
        p3bmv.BudgetModelCLIView(bmvm).cmdloop() # Application CLI loop

        # p3bm.execute_worklow_categorization(bm, "boa", p3bm.BM_WF_CATEGORIZATION)
        _ = "pause"
    except Exception as e:
        m = p3u.exc_msg(budmod, e)
        logger.error(m)
        raise
#endregion budmod() function
# ---------------------------------------------------------------------------- +
#region Local __main__ stand-alone
if __name__ == "__main__":
    try:

        configure_logging(settings.app_name)
        logger.setLevel(logging.DEBUG)
        # bm = p3fi.init_budget_model()  # How to load the budget model config?
        budmod() # Application Main()
    except Exception as e:
        m = p3u.exc_msg("__main__", e)
        logger.error(m)
    logger.info(f"Exiting {settings.app_name}...")
    exit(0)
#endregion Local __main__ stand-alone
# ---------------------------------------------------------------------------- +
#region tryit() function
def tryit() -> None:
    """Try something."""
    try:
        result = inspect.stack()[1][3]
        print(f"result: '{result}'")
    except Exception as e:
        print(f"Error: tryit(): {str(e)}")
#endregion tryit() function
# ---------------------------------------------------------------------------- +
